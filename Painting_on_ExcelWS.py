import sys
import openpyxl
from openpyxl.styles import PatternFill

args = sys.argv[1]

f = open(args, "r")
wb = openpyxl.Workbook()
ws = wb.active
lines = f.readlines()

for line in lines:
    coord_color = line.split(",")

    i_row = int(coord_color[1]) + 1
    i_col = int(coord_color[0].replace("(","")) + 1

    h_color = coord_color[2].replace(" #","").replace(")","")
    f_color = PatternFill(start_color=h_color, end_color=h_color, fill_type="solid")
    ws.cell(row=i_row, column=i_col).fill = f_color

wb.save("result.xlsx")
f.close
